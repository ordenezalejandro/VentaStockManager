# app/management/commands/importar_articulos.py

import argparse
import openpyxl
from datetime import datetime, timedelta
from django.core.management.base import BaseCommand
from articulo.models import Articulo
import traceback
import decimal
import os
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from openpyxl import Workbook

SERVICE_ACCOUNT_FILE = 'credentials.json'
SCOPES = ['https://www.googleapis.com/auth/spreadsheets.readonly']
FILE_ID = '1Zv9TDVJRDG_Ar-U4qTvlTcTiJ7RUpZnawxGwPpL4IZI'

def login_google_sheets():
    credentials = service_account.Credentials.from_service_account_file(
        SERVICE_ACCOUNT_FILE, scopes=SCOPES)
    return build('sheets', 'v4', credentials=credentials)

def download_sheet_from_google_sheets(sheet_id, range_name, ruta_descarga):
    service = login_google_sheets()
    sheet = service.spreadsheets()
    result = sheet.values().get(spreadsheetId=sheet_id, range=range_name).execute()
    values = result.get('values', [])

    if not values:
        print('No data found.')
        return None

    nombre_archivo = f'{sheet_id}.xlsx'
    ruta_archivo = os.path.join(ruta_descarga, nombre_archivo)

    if not os.path.exists(ruta_descarga):
        os.makedirs(ruta_descarga)

    wb = Workbook()
    ws = wb.active

    for row in values:
        ws.append(row)

    wb.save(ruta_archivo)
    return ruta_archivo

def buscar_y_cargar_documento():
    # ID del documento y rango de datos
    sheet_id = '1Zv9TDVJRDG_Ar-U4qTvlTcTiJ7RUpZnawxGwPpL4IZI'
    range_name = 'articulos!A1:Z1000'  # Ajusta el rango según tus necesidades

    ruta_documento = download_sheet_from_google_sheets(sheet_id, range_name, 'articulo.xlsx')

    if ruta_documento and os.path.exists(ruta_documento):
        call_command('cargar_articulo_xlsx', '-ruta_archivo', ruta_documento)
    else:
        print(f"Documento no encontrado en {ruta_documento}")

def generar_diccionario_letras_a_enteros():
    letras = "ABCDEFGHIJKLMNÑOPQRSTUVWXYZ"
    diccionario = {letra: indice + 1 for indice, letra in enumerate(letras)}
    return diccionario

DICCIONARIO_DE_LETRAS = generar_diccionario_letras_a_enteros()

class Command(BaseCommand):
    help = 'Importa datos desde un archivo .xlsx y crea instancias de Articulo'

    def add_arguments(self, parser):
        parser.add_argument('-ruta_archivo', type=str, help='Ruta al archivo .xlsx')

    def handle(self, *args, **kwargs):
        ruta_archivo = kwargs['ruta_archivo']

        self.stdout.write(self.style.SUCCESS(f'Importando datos desde {ruta_archivo}...'))

        try:
            if not ruta_archivo:
                ruta_archivo = download_sheet_from_google_sheets(FILE_ID, 'articulos!A1:Z1000', 'articulo.xlsx')
            if not ruta_archivo:
                raise ValueError("No se pudo descargar el archivo.")
            self.procesar_archivo_xlsx(ruta_archivo)
            self.stdout.write(self.style.SUCCESS('Importación completada exitosamente.'))
        except Exception as e:

            self.stdout.write(self.style.ERROR(f'Error al importar datos: {str(e)}'))
            self.stdout.write(self.style.ERROR(traceback.format_exc()))


    def procesar_archivo_xlsx(self, ruta_archivo):
        wb = openpyxl.load_workbook(ruta_archivo)
        sheet = wb.active
        for i, row in enumerate(sheet.iter_rows(min_row=4, values_only=True)):
            if not row[0] or not row[1] or not row[3] or not row[3] or (row[3]  is str and row[3].replace('$','') == ''):  # Si la primera celda está vacía, saltar la
                if row[1]:
                    self.stdout.write(f'esta fila no se proceso fila {i+4} row {row[0]} {row[1]}')
                continue
            try:
                nombre = row[0]
                codigo_interno = row[1]
                letra = DICCIONARIO_DE_LETRAS[row[1][0].upper()]
                codigo =f'{letra}{row[1][1:]}'
                try:
                    precio_minorista = decimal.Decimal(row[3].replace('$', ''))
                except Exception as e:
                    self.stdout.write(self.style.ERROR(f'Error al procesar la fila {i+1}: {str(e)} , se pondra a 0'))
                    precio_minorista = 0
                precio_mayorista = round(precio_minorista*decimal.Decimal(0.97), 2)

                if not all([nombre, codigo_interno, precio_minorista]):
                    self.stdout.write(self.style.ERROR(f'Fila {i+1} no se agregó: nombre, código interno o precio minorista es None.'))            
            except Exception as e:

                self.stdout.write(self.style.ERROR(f'Error al procesar la fila {i+1}: {str(e)}'))
                self.stdout.write(self.style.ERROR(traceback.format_exc()))
                continue

           
                
            # codigo_interno = self.generar_codigo_interno(nombre)

            articulo, creado = Articulo.objects.get_or_create(
                codigo_interno=codigo_interno,
                nombre=nombre,
                defaults={
                'codigo': codigo,
                'stock': 100,
                'vencimiento': datetime.now() + timedelta(days=90)
                }
            )
            articulo.precio_minorista = precio_minorista
            articulo.precio_mayorista = precio_mayorista
            articulo.stock = 100
            if articulo.codigo is None:
                articulo.codigo = codigo_interno
            articulo.stock = 100
            articulo.save()
            
    def generar_codigo_interno(self, nombre):
        primera_letra = nombre[0].lower()

        if Articulo.objects.filter(codigo_interno__startswith=primera_letra).exists():
            codigo_interno = nombre[:2].lower()
        else:
            codigo_interno = primera_letra

        return codigo_interno
